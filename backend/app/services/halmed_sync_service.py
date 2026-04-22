"""HZZO drug list sync service.

Downloads the official Osnovna + Dopunska lista lijekova from HZZO (.xlsx),
parses all drug entries, and upserts into the local drug_list table.

Data source: https://hzzo.hr/zdravstvena-zastita/objavljene-liste-lijekova
Updated monthly by HZZO, no auth required.

Columns in HZZO .xlsx:
  ATK šifra, Oznaka ograničenja primjene, Nezaštićeni naziv lijeka (INN),
  Način primjene, Nositelj odobrenja, Zaštićeni naziv lijeka (brand),
  Oblik/jačina i pakiranje, [Doplata], R/RS, PSL, Oznaka indikacije,
  Oznaka smjernice, Stopa PDV-a
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from datetime import UTC, datetime

import httpx
import openpyxl
from sqlalchemy import func, select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.database import async_session
from app.models.drug_list import DrugListItem

logger = logging.getLogger(__name__)

# HZZO drug list URLs (latest — March 2026)
# These URLs change when HZZO publishes new lists.
# The sync discovers the latest URLs by scraping the HZZO page.
HZZO_LISTE_URL = "https://hzzo.hr/zdravstvena-zastita/objavljene-liste-lijekova"

# Known direct URLs (fallback if page scrape fails)
HZZO_OSNOVNA_URL = (
    "https://hzzo.hr/sites/default/files/_web_OLL%20po%20dijelovima_stupa%20na%20snagu%20_16_03_2026..xlsx"
)
HZZO_DOPUNSKA_URL = (
    "https://hzzo.hr/sites/default/files/_web_DLL%20po%20dijelovima_stupa%20na%20snagu%20_16_03_2026._0.xlsx"
)

# Fallback URL expiration check — prevents using outdated drug data
HZZO_FALLBACK_URLS_DATE = datetime(2026, 3, 16)
HZZO_FALLBACK_MAX_AGE_MONTHS = 6

# Sheets that contain drug data (prefix match)
DRUG_SHEET_PREFIXES = ("OLL-", "DLL-", "OL-magistralni", "DL-magistralni")


def _cell(row: tuple, idx: int) -> str:
    """Extract cell value as stripped string."""
    val = row[idx] if idx < len(row) else None
    return str(val).strip() if val is not None else ""


def _parse_atk(raw: str) -> tuple[str, str]:
    """Parse HZZO 'ATK šifra' like 'A01AB12 451' into (atk, hzzo_sifra)."""
    raw = raw.strip()
    if not raw:
        return "", ""
    # Split on whitespace — first part is ATC code, rest is HZZO internal code
    parts = raw.split(None, 1)
    atk = parts[0] if parts else ""
    hzzo_sifra = parts[1] if len(parts) > 1 else ""
    return atk, hzzo_sifra


def _extract_jacina(oblik: str) -> str:
    """Try to extract strength from form/strength/packaging string.

    E.g. 'tbl. 30x20 mg' → '20 mg', 'gel oral. 2%, 1x40 g' → '2%'
    """
    # Look for patterns like '20 mg', '500mg', '2%', '1000 mg'
    m = re.search(r"(\d+(?:[.,]\d+)?\s*(?:mg|g|ml|%|IU|mL|mcg|µg))", oblik, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""


def _parse_drug_row(row: tuple, hzzo_lista: str, has_doplata: bool) -> dict | None:
    """Parse a single drug row from HZZO .xlsx into a drug dict."""
    # Column mapping (0-indexed):
    # 0: ATK šifra
    # 1: Oznaka ograničenja primjene
    # 2: Nezaštićeni naziv lijeka (INN)
    # 3: Način primjene
    # 4: Nositelj odobrenja
    # 5: Zaštićeni naziv lijeka (brand)
    # 6: Oblik, jačina i pakiranje
    # 7: Doplata (only in DLL) OR R/RS (in OLL without doplata)
    # 8+: R/RS, PSL, etc.

    atk_raw = _cell(row, 0)
    if not atk_raw:
        return None

    atk, hzzo_sifra = _parse_atk(atk_raw)
    inn_name = _cell(row, 2)  # Nezaštićeni naziv (generic)
    nacin = _cell(row, 3)  # Način primjene
    nositelj = _cell(row, 4)  # Nositelj odobrenja
    brand = _cell(row, 5)  # Zaštićeni naziv (brand)
    oblik = _cell(row, 6)  # Oblik, jačina i pakiranje

    # Determine R/RS and Doplata based on whether the sheet has doplata column
    doplata = ""
    r_rs = ""
    if has_doplata:
        doplata = _cell(row, 7)
        r_rs = _cell(row, 8)
    else:
        r_rs = _cell(row, 7)

    # Use brand name as naziv; if empty, use INN
    naziv = brand if brand else inn_name
    if not naziv:
        return None  # Skip rows with no name at all

    jacina = _extract_jacina(oblik)

    # Build search text from all relevant fields
    search = f"{naziv} {inn_name} {oblik} {atk} {nositelj} {hzzo_sifra}".lower()

    return {
        "atk": atk,
        "naziv": naziv,
        "oblik": oblik,
        "jacina": jacina,
        "inn": inn_name,
        "nositelj_odobrenja": nositelj,
        "hzzo_sifra": hzzo_sifra,
        "hzzo_lista": hzzo_lista,
        "r_rs": r_rs[:3],  # Max 3 chars for safety
        "nacin_primjene": nacin[:5],
        "doplata": doplata,
        "aktivan": True,
        "search_text": search,
    }


def _parse_xlsx(data: bytes, hzzo_lista: str) -> list[dict]:
    """Parse HZZO .xlsx file and return list of drug dicts."""
    drugs: list[dict] = []
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        # Only process drug data sheets
        if not any(sheet_name.startswith(p) for p in DRUG_SHEET_PREFIXES):
            continue

        ws = wb[sheet_name]
        has_doplata = "DLL" in sheet_name.upper() or "DL-" in sheet_name.lower()

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Skip header row(s) — first row is always headers
            if i == 0:
                continue
            # Skip empty rows
            if not row or not row[0]:
                continue

            drug = _parse_drug_row(row, hzzo_lista, has_doplata)
            if drug:
                drugs.append(drug)

    wb.close()
    return drugs


async def _discover_hzzo_urls() -> tuple[str, str]:
    """Scrape HZZO page to find latest .xlsx download URLs.

    Falls back to hardcoded URLs if discovery fails, but only if they're
    not too old (see HZZO_FALLBACK_MAX_AGE_MONTHS).
    """
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            resp = await client.get(HZZO_LISTE_URL)
            if resp.status_code != 200:
                raise httpx.HTTPStatusError("non-200", request=resp.request, response=resp)

            import re

            html = resp.text
            # Find .xlsx links — Osnovna and Dopunska
            xlsx_links = re.findall(r'href="([^"]*\.xlsx[^"]*)"', html)
            osnovna = ""
            dopunska = ""
            for link in xlsx_links:
                lower = link.lower()
                if "oll" in lower and not osnovna:
                    osnovna = link if link.startswith("http") else f"https://hzzo.hr{link}"
                elif "dll" in lower and not dopunska:
                    dopunska = link if link.startswith("http") else f"https://hzzo.hr{link}"

            if osnovna and dopunska:
                logger.info("HZZO URLs discovered: OLL=%s, DLL=%s", osnovna[:80], dopunska[:80])
                return osnovna, dopunska

            raise ValueError(f"Could not discover URLs from HZZO page (found {len(xlsx_links)} links)")
    except Exception as e:
        # Check if fallback is still usable

        fallback_age_days = (datetime.now(UTC).replace(tzinfo=UTC) - HZZO_FALLBACK_URLS_DATE.replace(tzinfo=UTC)).days
        age_months = fallback_age_days / 30

        if age_months > HZZO_FALLBACK_MAX_AGE_MONTHS:
            from app.services.cezih.exceptions import CezihError

            raise CezihError(
                f"HZZO URL discovery failed i fallback URL-ovi su stari {age_months:.0f} mjeseci. "
                f"Ažurirajte HZZO_OSNOVNA_URL i HZZO_DOPUNSKA_URL u halmed_sync_service.py. "
                f"Original error: {e}"
            ) from e

        logger.warning("HZZO URL discovery failed (%s), using fallback URLs (age=%.1f months)", e, age_months)
        return HZZO_OSNOVNA_URL, HZZO_DOPUNSKA_URL


async def sync_hzzo_drugs() -> dict:
    """Full sync: download HZZO lists, parse, and upsert into DB.

    Returns summary dict with counts.
    """
    logger.info("HZZO drug sync starting...")
    all_drugs: list[dict] = []

    # Discover latest URLs from HZZO page
    osnovna_url, dopunska_url = await _discover_hzzo_urls()

    async with httpx.AsyncClient(timeout=120, follow_redirects=True) as client:
        # Download Osnovna lista
        try:
            resp = await client.get(osnovna_url)
            if resp.status_code == 200 and resp.content:
                drugs_oll = _parse_xlsx(resp.content, "OLL")
                logger.info("HZZO Osnovna lista: %d drugs", len(drugs_oll))
                all_drugs.extend(drugs_oll)
            else:
                logger.warning("HZZO Osnovna download failed: HTTP %d", resp.status_code)
        except Exception as e:
            logger.error("HZZO Osnovna download error: %s", e)

        await asyncio.sleep(1)

        # Download Dopunska lista
        try:
            resp = await client.get(dopunska_url)
            if resp.status_code == 200 and resp.content:
                drugs_dll = _parse_xlsx(resp.content, "DLL")
                logger.info("HZZO Dopunska lista: %d drugs", len(drugs_dll))
                all_drugs.extend(drugs_dll)
            else:
                logger.warning("HZZO Dopunska download failed: HTTP %d", resp.status_code)
        except Exception as e:
            logger.error("HZZO Dopunska download error: %s", e)

    if not all_drugs:
        logger.warning("HZZO sync: no drugs fetched, skipping DB update")
        return {"fetched": 0, "upserted": 0, "error": "No drugs fetched from HZZO"}

    # Deduplicate by atk+naziv+oblik (keep last occurrence)
    seen: dict[str, dict] = {}
    for drug in all_drugs:
        key = f"{drug['atk']}|{drug['naziv']}|{drug['oblik']}"
        seen[key] = drug
    all_drugs = list(seen.values())

    logger.info("HZZO sync: %d unique drug entries, upserting...", len(all_drugs))

    now = datetime.now(UTC)
    upserted = 0

    async with async_session() as db:
        # Mark all existing as inactive, then re-activate what we get
        await db.execute(text("UPDATE drug_list SET aktivan = false"))

        # Upsert in batches
        batch_size = 500
        for i in range(0, len(all_drugs), batch_size):
            batch = all_drugs[i : i + batch_size]
            for drug in batch:
                drug["synced_at"] = now

            stmt = pg_insert(DrugListItem).values(batch)
            stmt = stmt.on_conflict_do_update(
                constraint="uq_drug_list_atk_naziv_oblik",
                set_={
                    "atk": stmt.excluded.atk,
                    "naziv": stmt.excluded.naziv,
                    "oblik": stmt.excluded.oblik,
                    "jacina": stmt.excluded.jacina,
                    "inn": stmt.excluded.inn,
                    "nositelj_odobrenja": stmt.excluded.nositelj_odobrenja,
                    "hzzo_sifra": stmt.excluded.hzzo_sifra,
                    "hzzo_lista": stmt.excluded.hzzo_lista,
                    "r_rs": stmt.excluded.r_rs,
                    "nacin_primjene": stmt.excluded.nacin_primjene,
                    "doplata": stmt.excluded.doplata,
                    "aktivan": stmt.excluded.aktivan,
                    "search_text": stmt.excluded.search_text,
                    "synced_at": stmt.excluded.synced_at,
                },
            )
            await db.execute(stmt)
            upserted += len(batch)

        await db.commit()

    logger.info("HZZO sync complete: %d entries upserted", upserted)
    return {"fetched": len(all_drugs), "upserted": upserted}


async def search_drugs_db(query: str, limit: int = 20) -> list[dict]:
    """Search local drug_list table. Returns empty list if no data (caller falls back to mock)."""
    if not query or len(query) < 2:
        return []

    async with async_session() as db:
        # Check if we have any data
        count_result = await db.execute(select(func.count()).select_from(DrugListItem))
        count = count_result.scalar() or 0

        if count == 0:
            return []  # Caller should fall back to mock

        q = f"%{query.lower()}%"
        result = await db.execute(
            select(DrugListItem)
            .where(DrugListItem.aktivan.is_(True), DrugListItem.search_text.ilike(q))
            .order_by(DrugListItem.naziv)
            .limit(limit)
        )
        rows = result.scalars().all()
        return [
            {
                "atk": r.atk,
                "naziv": r.naziv,
                "oblik": r.oblik,
                "jacina": r.jacina,
            }
            for r in rows
        ]


# --- Scheduler ---

_sync_task: asyncio.Task | None = None


async def _is_sync_fresh() -> bool:
    """Check if drug data was synced within the last 7 days."""
    from app.models.drug_list import DrugListItem

    async with async_session() as db:
        result = await db.execute(
            select(func.max(DrugListItem.synced_at)).where(DrugListItem.aktivan == True)  # noqa: E712
        )
        last_sync = result.scalar()
        if last_sync is None:
            return False
        # last_sync is a datetime from the DB, but mypy sees it as str
        age = datetime.now(UTC) - last_sync  # type: ignore[operator]
        return age.days < 7


async def _sync_loop():
    """Run sync immediately on startup (if stale), then every 7 days."""
    await asyncio.sleep(5)

    # Skip initial sync if DB already has fresh data (< 7 days old)
    try:
        if await _is_sync_fresh():
            logger.info("HZZO initial sync skipped — data is fresh (< 7 days)")
        else:
            result = await sync_hzzo_drugs()
            logger.info("HZZO initial sync result: %s", result)
    except Exception:
        logger.exception("HZZO initial sync failed")

    # Weekly loop
    while True:
        await asyncio.sleep(7 * 24 * 3600)  # 7 days
        try:
            result = await sync_hzzo_drugs()
            logger.info("HZZO weekly sync result: %s", result)
        except Exception:
            logger.exception("HZZO weekly sync failed")


def start_sync_scheduler():
    """Start the background sync loop. Call once from app lifespan."""
    global _sync_task
    _sync_task = asyncio.create_task(_sync_loop())
    logger.info("HZZO drug sync scheduler started")


def stop_sync_scheduler():
    """Cancel the background sync loop."""
    global _sync_task
    if _sync_task and not _sync_task.done():
        _sync_task.cancel()
        logger.info("HZZO drug sync scheduler stopped")
